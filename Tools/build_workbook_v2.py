import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

OUT = sys.argv[1] if len(sys.argv) > 1 else "policy-sludge-scoring-workbook-v2.xlsx"

NAVY="1F3864"; BLUE="2E5496"; LIGHT="D9E1F2"; GREY="F2F2F2"; YEL="FFF2CC"; GREEN="E2EFDA"; WHITE="FFFFFF"; SUB="FBE4D5"
def font(sz=10,b=False,color="000000"): return Font(name="Arial",size=sz,bold=b,color=color)
thin=Side(style="thin",color="BFBFBF"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
wrap=Alignment(wrap_text=True,vertical="top"); center=Alignment(horizontal="center",vertical="center")
ctr_wrap=Alignment(horizontal="center",vertical="center",wrap_text=True); vcenter=Alignment(vertical="center")

wb=Workbook()

# ---------------- Instructions ----------------
ws=wb.active; ws.title="Instructions"; ws.sheet_view.showGridLines=False
ws.column_dimensions['A'].width=3; ws.column_dimensions['B'].width=112
def title(ws,cell,text): ws[cell]=text; ws[cell].font=font(16,True,WHITE); ws[cell].fill=PatternFill("solid",start_color=NAVY); ws[cell].alignment=vcenter
def h(ws,cell,text): ws[cell]=text; ws[cell].font=font(12,True,NAVY)
def p(ws,cell,text): ws[cell]=text; ws[cell].font=font(10); ws[cell].alignment=wrap
title(ws,'B2',"Policy-Sludge Output Evaluation — Scoring Workbook (v2, merged)"); ws.row_dimensions[2].height=26
rows=[
 ('B4','h',"What this workbook is"),
 ('B5','p',"Scoring instrument for the v2 Policy-Sludge Output Evaluation Rubric. Works for any AI output that identifies policy sludge (duplicative, overlapping, redundant, obsolete, ambiguous, or disproportionately burdensome requirements), not only the sample banking-board prompt. Pair with the Rubric v2, the LLM-as-Judge Prompt v2, and the Reviewer Protocol v2."),
 ('B7','h',"What changed in v2"),
 ('B8','p',"1) Integrity-weighted scoring (no longer equal): D1 20, D2 25, D3 20, D4 15, D5 10, D6 10 — integrity axes carry 55 of 100.  2) Scored subcriteria: actor mapping (D1/D5), binding-vs-guidance (D3), reform usefulness (D4).  3) Two added flags: F6 scope evasion, F7 unsupported central legal conclusion.  4) Graduated cap severity: F1/F3/F4 cap their dimension to 0 if Central, ≤1 if Peripheral."),
 ('B10','h',"Tabs"),
 ('B11','p',"• Scorecard — detailed assessment of ONE output. Enter six raw 0–4 scores, flags (with severity), and subcriteria checks; the sheet computes effective scores (after caps), the weighted overall /100, tier, and reliability overlay."),
 ('B12','p',"• Comparison — up to six outputs side by side. Enter each output's raw scores AND flags; the sheet computes effective scores, weighted overall, tier, overlay, and rank automatically (no manual cap math)."),
 ('B13','p',"• Citation Sample Log — record the citation spot-check (with central/peripheral tags) behind the integrity scores. Verdict counts feed D2/D3 and the flags."),
 ('B14','p',"• Rubric Reference — condensed anchors, weights, flags, severity, and tier bands."),
 ('B16','h',"Scoring math"),
 ('B17','p',"Each dimension scored 0–4 (raw) → caps applied → effective. Overall = Σ(weight × effective ÷ 4); weights sum to 100, so a perfect 4 across all dimensions = 100. Tiers: A 85–100 · B 70–84 · C 50–69 · D 30–49 · E 0–29. Reliability overlay (Integrity-flagged / Scope-flagged) is reported alongside the tier and governs reliance regardless of the number."),
 ('B19','h',"How to score (short version)"),
 ('B20','p',"1) Reconstruct the request spec and build a coverage map. 2) Mark coverage incl. every actor class → D1. 3) Spot-check a citation sample, tag central/peripheral → D2/D3 + flags. 4) Judge sludge analysis + reform levers → D4. 5) Judge structure/actor-mapping (D5) and honesty (D6). 6) Check the three subcriteria. 7) Enter raw scores + flags here; write the narrative. Anchor every score to a concrete quote."),
 ('B22','h',"Reading the result"),
 ('B23','p',"Example: \"Tier B (78/100) — INTEGRITY-FLAGGED (F1 central).\" Because integrity carries 55 of 100 and central integrity flags zero their dimension, a compromised output usually drops a tier or more AND carries the overlay."),
]
for cell,kind,text in rows:
    (h if kind=='h' else p)(ws,cell,text)
for r in (5,8,11,12,13,14,17,20,23):
    ws.row_dimensions[r].height={5:46,8:60,17:46,20:60,23:40}.get(r,30)

def hdrcell(ws,r,c,text,fillc=BLUE,sz=10):
    cc=ws.cell(row=r,column=c,value=text); cc.font=font(sz,True,WHITE); cc.fill=PatternFill("solid",start_color=fillc); cc.alignment=ctr_wrap; cc.border=border; return cc
def lbl(ws,cell,t): ws[cell]=t; ws[cell].font=font(10,True,NAVY)
def inp(ws,cell,fillc=YEL): ws[cell].fill=PatternFill("solid",start_color=fillc); ws[cell].border=border; ws[cell].font=font(10); ws[cell].alignment=wrap

# ---------------- Scorecard ----------------
sc=wb.create_sheet("Scorecard"); sc.sheet_view.showGridLines=False
for c,w in {'A':36,'B':10,'C':10,'D':22,'E':10,'F':12,'G':40,'H':34}.items(): sc.column_dimensions[c].width=w
sc.merge_cells('A1:H1'); sc['A1']="Policy-Sludge Output — Scorecard (v2)"; sc['A1'].font=font(15,True,WHITE); sc['A1'].fill=PatternFill("solid",start_color=NAVY); sc['A1'].alignment=vcenter; sc.row_dimensions[1].height=24
lbl(sc,'A2',"Output ID:"); inp(sc,'B2'); sc.merge_cells('B2:D2')
lbl(sc,'E2',"Reviewer:"); inp(sc,'F2'); sc.merge_cells('F2:H2')
lbl(sc,'A3',"Date:"); inp(sc,'B3'); sc.merge_cells('B3:D3')
lbl(sc,'E3',"Run mode (human/LLM):"); inp(sc,'F3'); sc.merge_cells('F3:H3')
lbl(sc,'A4',"Original request / context:"); inp(sc,'B4'); sc.merge_cells('B4:H4'); sc.row_dimensions[4].height=30

hdr_r=6
for i,htxt in enumerate(["Dimension","Weight","Raw (0–4)","Cap applied","Effective","Weighted pts","Evidence quote (≤25 words)","Notes / rationale"]):
    hdrcell(sc,hdr_r,1+i,htxt)
dims=[("D1  Coverage & Comprehensiveness",20),("D2  Source & Citation Integrity",25),("D3  Substantive Accuracy & Currency",20),("D4  Sludge-Analysis Quality",15),("D5  Structure, Granularity & Fitness",10),("D6  Calibration & Transparency",10)]
first=hdr_r+1  # 7
for i,(d,wt) in enumerate(dims):
    r=first+i
    sc.cell(row=r,column=1,value=d).font=font(10,True); sc.cell(row=r,column=1).alignment=wrap; sc.cell(row=r,column=1).border=border
    wc=sc.cell(row=r,column=2,value=wt); wc.font=font(10,True); wc.alignment=center; wc.border=border; wc.fill=PatternFill("solid",start_color=GREY)
    raw=sc.cell(row=r,column=3); raw.fill=PatternFill("solid",start_color=YEL); raw.border=border; raw.alignment=center; raw.font=font(10,True)
    for col in (4,5,6,7,8):
        cc=sc.cell(row=r,column=col); cc.border=border; cc.alignment=(center if col in (5,6) else wrap); cc.font=font(10,True if col in (5,6) else False)
        if col==4: cc.font=font(9); cc.alignment=center
    sc.row_dimensions[r].height=34
# rows: D1=7..D6=12
# flags table
fl_hdr=15
hdrcell(sc,fl_hdr,1,"Critical flag"); hdrcell(sc,fl_hdr,2,"Present? (Y/N)"); hdrcell(sc,fl_hdr,3,"Severity")
sc.merge_cells(start_row=fl_hdr,start_column=4,end_row=fl_hdr,end_column=8); hdrcell(sc,fl_hdr,4,"Evidence (specific failing citation / claim)")
flags=[("F1  Fabricated source / quotation",True),("F2  Misattribution",False),("F3  Stale-as-current",True),("F4  Invented obligation",True),("F5  False exhaustiveness",False),("F6  Material scope evasion",False),("F7  Unsupported central legal conclusion",False)]
fl_first=fl_hdr+1  # 16
for i,(f,has_sev) in enumerate(flags):
    r=fl_first+i
    sc.cell(row=r,column=1,value=f).font=font(10); sc.cell(row=r,column=1).alignment=wrap; sc.cell(row=r,column=1).border=border
    pc=sc.cell(row=r,column=2,value="N"); pc.fill=PatternFill("solid",start_color=YEL); pc.border=border; pc.alignment=center
    sv=sc.cell(row=r,column=3); sv.border=border; sv.alignment=center
    if has_sev: sv.value="Central"; sv.fill=PatternFill("solid",start_color=YEL)
    else: sv.value="—"; sv.fill=PatternFill("solid",start_color=GREY); sv.font=font(9)
    sc.merge_cells(start_row=r,start_column=4,end_row=r,end_column=8)
    ec=sc.cell(row=r,column=4); ec.border=border; ec.alignment=wrap
    sc.row_dimensions[r].height=24
# F1=16,F2=17,F3=18,F4=19,F5=20,F6=21,F7=22
# effective + cap-applied formulas
sc['E7']='=IF($B$21="Y",MIN(C7,1),C7)'
sc['D7']='=IF($B$21="Y","capped ≤1 (F6)","—")'
sc['E8']='=IF(AND($B$16="Y",$C$16="Central"),0,IF(OR($B$16="Y",$B$17="Y",$B$22="Y"),MIN(C8,1),C8))'
sc['D8']='=IF(AND($B$16="Y",$C$16="Central"),"capped 0 (F1 central)",IF(OR($B$16="Y",$B$17="Y",$B$22="Y"),"capped ≤1 (F1/F2/F7)","—"))'
sc['E9']='=IF(OR(AND($B$18="Y",$C$18="Central"),AND($B$19="Y",$C$19="Central")),0,IF(OR($B$18="Y",$B$19="Y"),MIN(C9,1),C9))'
sc['D9']='=IF(OR(AND($B$18="Y",$C$18="Central"),AND($B$19="Y",$C$19="Central")),"capped 0 (F3/F4 central)",IF(OR($B$18="Y",$B$19="Y"),"capped ≤1 (F3/F4)","—"))'
sc['E10']='=C10'; sc['D10']="—"
sc['E11']='=C11'; sc['D11']="—"
sc['E12']='=IF($B$20="Y",MIN(C12,1),C12)'; sc['D12']='=IF($B$20="Y","capped ≤1 (F5)","—")'
for r in range(7,13): sc.cell(row=r,column=6,value=f"=B{r}*E{r}/4")
# subcriteria
sub_hdr=24
sc.merge_cells(start_row=sub_hdr,start_column=1,end_row=sub_hdr,end_column=8)
hdrcell(sc,sub_hdr,1,"Subcriteria checks  (inform the RAW score per these caps — apply when entering raw)",fillc="C55A11")
subs=[("Actor mapping — every required actor class covered & duties tied to precise actors?","Miss a required actor class → cap D1 ≤2;  generic 'board oversight' language → cap D5 ≤2"),
      ("Binding-vs-guidance — binding law clearly separated from non-binding guidance?","Systematic guidance-as-law blending → cap D3 ≤2"),
      ("Reform usefulness — reform levers indicated (consolidate/delegate/clarify/keep)?","No reform framing → cap D4 ≤3;  purely descriptive list → cap D4 ≤2")]
for i,(q,note) in enumerate(subs):
    r=sub_hdr+1+i
    sc.cell(row=r,column=1,value=q).font=font(10); sc.cell(row=r,column=1).alignment=wrap; sc.cell(row=r,column=1).border=border
    yn=sc.cell(row=r,column=2,value="Y"); yn.fill=PatternFill("solid",start_color=YEL); yn.border=border; yn.alignment=center
    sc.merge_cells(start_row=r,start_column=3,end_row=r,end_column=8)
    nc=sc.cell(row=r,column=3,value=note); nc.font=font(9,color="833C00"); nc.alignment=wrap; nc.border=border
    sc.row_dimensions[r].height=26
# subcriteria rows 25,26,27
# results
res=29
sc.cell(row=res,column=1,value="Overall score (/100)").font=font(11,True,NAVY)
oc=sc.cell(row=res,column=2,value="=ROUND(SUM(F7:F12),0)"); oc.font=font(14,True); oc.fill=PatternFill("solid",start_color=GREEN); oc.alignment=center; oc.border=border
sc.merge_cells(start_row=res,start_column=2,end_row=res,end_column=3)
sc.cell(row=res+1,column=1,value="Tier").font=font(11,True,NAVY)
tier='=IF(B29>=85,"A — Publication-ready",IF(B29>=70,"B — Strong",IF(B29>=50,"C — Usable w/ revision",IF(B29>=30,"D — Weak","E — Inadequate"))))'
tc=sc.cell(row=res+1,column=2,value=tier); tc.font=font(11,True); tc.alignment=vcenter; tc.border=border; sc.merge_cells(start_row=res+1,start_column=2,end_row=res+1,end_column=6)
sc.cell(row=res+2,column=1,value="Reliability overlay").font=font(11,True,NAVY)
overlay='=IF(OR(B16="Y",B17="Y",B18="Y",B19="Y",B20="Y",B22="Y"),IF(B21="Y","INTEGRITY + SCOPE-FLAGGED — verify before reliance","INTEGRITY-FLAGGED — verify before reliance"),IF(B21="Y","SCOPE-FLAGGED — verify scope","Clean"))'
ov=sc.cell(row=res+2,column=2,value=overlay); ov.font=font(11,True); ov.alignment=vcenter; ov.border=border; sc.merge_cells(start_row=res+2,start_column=2,end_row=res+2,end_column=8)
sc.cell(row=res+3,column=1,value="Result line").font=font(11,True,NAVY)
resline='="Tier "&LEFT(B30,1)&" ("&B29&"/100) — "&B31'
rc=sc.cell(row=res+3,column=2,value=resline); rc.font=font(11,True); rc.alignment=vcenter; sc.merge_cells(start_row=res+3,start_column=2,end_row=res+3,end_column=8)
# narrative
nar=34
labels=[("Top strengths (3–5, each tied to a dimension + example)",54),("Top weaknesses (3–5, ordered by severity)",54),("Subcriteria verdicts (actor mapping / binding-vs-guidance / reform usefulness)",34),("Prioritized remediation (3–5 highest-value fixes)",46),("Reliance recommendation (one sentence)",26)]
r=nar
for lab,ht in labels:
    sc.cell(row=r,column=1,value=lab).font=font(10,True,NAVY); sc.cell(row=r,column=1).alignment=wrap
    sc.merge_cells(start_row=r,start_column=2,end_row=r,end_column=8)
    cc=sc.cell(row=r,column=2); cc.border=border; cc.alignment=wrap
    sc.row_dimensions[r].height=ht; r+=1
# validations
dv=DataValidation(type="list",formula1='"0,1,2,3,4"',allow_blank=True); sc.add_data_validation(dv); dv.add("C7:C12")
dvyn=DataValidation(type="list",formula1='"Y,N"',allow_blank=True); sc.add_data_validation(dvyn); dvyn.add("B16:B22"); dvyn.add("B25:B27")
dvsev=DataValidation(type="list",formula1='"Central,Peripheral"',allow_blank=True); sc.add_data_validation(dvsev); dvsev.add("C16"); dvsev.add("C18"); dvsev.add("C19")

# ---------------- Comparison ----------------
cp=wb.create_sheet("Comparison"); cp.sheet_view.showGridLines=False
cp.column_dimensions['A'].width=42
for col in ['B','C','D','E','F','G']: cp.column_dimensions[col].width=14
cp.merge_cells('A1:G1'); cp['A1']="Policy-Sludge Outputs — Comparison (up to 6, auto-computed)"; cp['A1'].font=font(15,True,WHITE); cp['A1'].fill=PatternFill("solid",start_color=NAVY); cp['A1'].alignment=vcenter; cp.row_dimensions[1].height=24
cp.merge_cells('A2:G2'); cp['A2']="Enter each output's six RAW 0–4 scores and its flags. The sheet applies caps and computes effective scores, weighted overall, tier, overlay, and rank."; cp['A2'].font=font(9,True); cp.row_dimensions[2].height=26
hdrcell(cp,3,1,"Row")
for i in range(6): hdrcell(cp,3,2+i,"Output %s"%chr(65+i))
def rowlabel(r,text,fillc=None,bold=False,sz=10):
    c=cp.cell(row=r,column=1,value=text); c.font=font(sz,bold,"000000" if not fillc else WHITE); c.alignment=wrap; c.border=border
    if fillc: c.fill=PatternFill("solid",start_color=fillc); c.font=font(sz,True,WHITE)
# raw scores
rawlabels=["D1 Coverage & Comprehensiveness (raw)","D2 Source & Citation Integrity (raw)","D3 Substantive Accuracy & Currency (raw)","D4 Sludge-Analysis Quality (raw)","D5 Structure & Fitness (raw)","D6 Calibration & Transparency (raw)"]
for i,l in enumerate(rawlabels):
    r=4+i; rowlabel(r,l)
    for col in range(2,8):
        cc=cp.cell(row=r,column=col); cc.fill=PatternFill("solid",start_color=YEL); cc.border=border; cc.alignment=center
# raw D1=4..D6=9
rowlabel(10,"FLAGS  (Y/N; severity where shown)",fillc=BLUE)
cp.merge_cells('B10:G10')
flagrows=[("F1 Fabricated (Y/N)",11,"yn"),("F1 severity (Central/Peripheral)",12,"sev"),("F2 Misattribution (Y/N)",13,"yn"),("F3 Stale-as-current (Y/N)",14,"yn"),("F3 severity",15,"sev"),("F4 Invented obligation (Y/N)",16,"yn"),("F4 severity",17,"sev"),("F5 False exhaustiveness (Y/N)",18,"yn"),("F6 Scope evasion (Y/N)",19,"yn"),("F7 Unsupported central conclusion (Y/N)",20,"yn")]
for label,r,kind in flagrows:
    rowlabel(r,label,sz=9)
    for col in range(2,8):
        cc=cp.cell(row=r,column=col); cc.fill=PatternFill("solid",start_color=YEL); cc.border=border; cc.alignment=center
        cc.value="N" if kind=="yn" else "Central"
        cc.font=font(9)
# effective (computed)
rowlabel(21,"EFFECTIVE (auto)",fillc=GREY); cp.merge_cells('B21:G21')
efflabels=["eff D1","eff D2","eff D3","eff D4","eff D5","eff D6"]
for i,l in enumerate(efflabels):
    r=22+i; rowlabel(r,l,sz=9)
cols=['B','C','D','E','F','G']
for col in cols:
    cp[f'{col}22']=f'=IF({col}19="Y",MIN({col}4,1),{col}4)'
    cp[f'{col}23']=f'=IF(AND({col}11="Y",{col}12="Central"),0,IF(OR({col}11="Y",{col}13="Y",{col}20="Y"),MIN({col}5,1),{col}5))'
    cp[f'{col}24']=f'=IF(OR(AND({col}14="Y",{col}15="Central"),AND({col}16="Y",{col}17="Central")),0,IF(OR({col}14="Y",{col}16="Y"),MIN({col}6,1),{col}6))'
    cp[f'{col}25']=f'={col}7'
    cp[f'{col}26']=f'={col}8'
    cp[f'{col}27']=f'=IF({col}18="Y",MIN({col}9,1),{col}9)'
    for rr in range(22,28):
        cp[f'{col}{rr}'].alignment=center; cp[f'{col}{rr}'].font=font(9); cp[f'{col}{rr}'].border=border
# results
rowlabel(29,"Weighted overall (/100)",bold=True); rowlabel(30,"Tier",bold=True); rowlabel(31,"Reliability overlay",bold=True); rowlabel(32,"Rank (by overall)",bold=True)
for col in cols:
    cp[f'{col}29']=f'=IF(COUNT({col}4:{col}9)<6,"",ROUND((20*{col}22+25*{col}23+20*{col}24+15*{col}25+10*{col}26+10*{col}27)/4,0))'
    cp[f'{col}29'].font=font(12,True); cp[f'{col}29'].fill=PatternFill("solid",start_color=GREEN); cp[f'{col}29'].alignment=center; cp[f'{col}29'].border=border
    cp[f'{col}30']=f'=IF({col}29="","",IF({col}29>=85,"A",IF({col}29>=70,"B",IF({col}29>=50,"C",IF({col}29>=30,"D","E")))))'
    cp[f'{col}31']=f'=IF({col}29="","",IF(OR({col}11="Y",{col}13="Y",{col}14="Y",{col}16="Y",{col}18="Y",{col}20="Y"),IF({col}19="Y","INTEG+SCOPE","INTEGRITY"),IF({col}19="Y","SCOPE","Clean")))'
    cp[f'{col}32']=f'=IF({col}29="","",SUMPRODUCT(($B$29:$G$29<>"")*($B$29:$G$29>{col}29))+1)'
    for rr in (30,31,32):
        cp[f'{col}{rr}'].font=font(10,True); cp[f'{col}{rr}'].alignment=center; cp[f'{col}{rr}'].border=border
cp.merge_cells('A34:G34'); cp['A34']="Note: Rank is by numeric overall only. Any output marked INTEGRITY / INTEG+SCOPE in the overlay row should not be relied on regardless of rank."; cp['A34'].font=font(9,True); cp['A34'].alignment=wrap; cp.row_dimensions[34].height=28
# validations
dv2=DataValidation(type="list",formula1='"0,1,2,3,4"',allow_blank=True); cp.add_data_validation(dv2); dv2.add("B4:G9")
ynrows=[11,13,14,16,18,19,20]; sevrows=[12,15,17]
dvyn2=DataValidation(type="list",formula1='"Y,N"',allow_blank=True); cp.add_data_validation(dvyn2)
for r in ynrows: dvyn2.add(f"B{r}:G{r}")
dvsev2=DataValidation(type="list",formula1='"Central,Peripheral"',allow_blank=True); cp.add_data_validation(dvsev2)
for r in sevrows: dvsev2.add(f"B{r}:G{r}")

# ---------------- Citation Sample Log ----------------
cl=wb.create_sheet("Citation Sample Log"); cl.sheet_view.showGridLines=False
cl.merge_cells('A1:I1'); cl['A1']="Citation Spot-Check Log"; cl['A1'].font=font(15,True,WHITE); cl['A1'].fill=PatternFill("solid",start_color=NAVY); cl['A1'].alignment=vcenter; cl.row_dimensions[1].height=24
verdicts=["SUPPORTS","WRONG-ATTRIBUTION","NOT-SUPPORTED","STALE","FABRICATED","UNVERIFIABLE"]
cl['A3']="Verdict counts:"; cl['A3'].font=font(10,True,NAVY)
for i,v in enumerate(verdicts):
    cl.cell(row=3,column=2+i,value=v).font=font(8,True); cl.cell(row=3,column=2+i).alignment=ctr_wrap; cl.cell(row=3,column=2+i).fill=PatternFill("solid",start_color=GREY)
    cl.cell(row=4,column=2+i,value=f'=COUNTIF($F$7:$F$56,"{v}")').alignment=center; cl.cell(row=4,column=2+i).font=font(10,True)
heads=["#","Section in output","Citation as stated","Claim it supports","Verified against (source / URL)","Verdict","Central / Peripheral","Reason (1 line)","Reviewer"]
widths=[5,18,26,28,28,17,15,30,11]
hr=6
for i,(htxt,w) in enumerate(zip(heads,widths)):
    hdrcell(cl,hr,1+i,htxt); cl.column_dimensions[chr(65+i)].width=w
for r in range(hr+1,hr+51):
    cl.cell(row=r,column=1,value=r-hr).font=font(9); cl.cell(row=r,column=1).alignment=center
    for col in range(1,10):
        cc=cl.cell(row=r,column=col); cc.border=border; cc.alignment=wrap; cc.font=font(9)
        if col in (2,3,4,5,6,7,8,9): cc.fill=PatternFill("solid",start_color=YEL)
dvv=DataValidation(type="list",formula1='"%s"'%",".join(verdicts),allow_blank=True); cl.add_data_validation(dvv); dvv.add(f"F{hr+1}:F{hr+50}")
dvc=DataValidation(type="list",formula1='"Central,Peripheral,n/a"',allow_blank=True); cl.add_data_validation(dvc); dvc.add(f"G{hr+1}:G{hr+50}")

# ---------------- Rubric Reference ----------------
rf=wb.create_sheet("Rubric Reference"); rf.sheet_view.showGridLines=False
rf.column_dimensions['A'].width=40; rf.column_dimensions['B'].width=12; rf.column_dimensions['C'].width=84
rf.merge_cells('A1:C1'); rf['A1']="Rubric Reference (v2) — anchors, weights, flags, tiers"; rf['A1'].font=font(14,True,WHITE); rf['A1'].fill=PatternFill("solid",start_color=NAVY); rf['A1'].alignment=vcenter; rf.row_dimensions[1].height=22
def rh(cell,t): rf[cell]=t; rf[cell].font=font(11,True,NAVY)
rh('A3',"Dimensions (score 0–4) and weights")
hdrcell(rf,4,1,"Dimension"); hdrcell(rf,4,2,"Weight"); hdrcell(rf,4,3,"What 4 vs 2 vs 0 looks like / subcriterion")
dim_desc=[("D1 Coverage & Comprehensiveness","20","Fills the coverage map across all axes/sources/areas; even depth. 4=near-complete; 2=noticeable gaps; 0=wrong scope. Sub: missing a required actor class → cap ≤2."),
 ("D2 Source & Citation Integrity","25","Citations real, correctly attributed, precise, faithful. 4=specific & verifiable; 2=imprecise/over-broad; 1=a misattribution/unsupported claim; 0=fabrication."),
 ("D3 Substantive Accuracy & Currency","20","Correct + in force now. 4=correct & current; 2=some mischaracterization/currency error; 0=stale-as-current. Sub: guidance-as-law blending → cap ≤2."),
 ("D4 Sludge-Analysis Quality","15","Maps duplication/overlap/obsolescence; discriminates true sludge; few false positives. Sub: no reform levers → cap ≤3; pure list → ≤2."),
 ("D5 Structure, Granularity & Fitness","10","Organized, granular, usable for the audience. Sub: generic 'board oversight' actor language → cap ≤2."),
 ("D6 Calibration & Transparency","10","Honest about scope/gaps/uncertainty/method; no overclaiming. 0=affirmatively claims unearned exhaustiveness.")]
r=5
for a,w,b in dim_desc:
    rf.cell(row=r,column=1,value=a).font=font(10,True); rf.cell(row=r,column=1).alignment=wrap; rf.cell(row=r,column=1).border=border
    rf.cell(row=r,column=2,value=w).font=font(10,True); rf.cell(row=r,column=2).alignment=center; rf.cell(row=r,column=2).border=border
    rf.cell(row=r,column=3,value=b).font=font(10); rf.cell(row=r,column=3).alignment=wrap; rf.cell(row=r,column=3).border=border
    rf.row_dimensions[r].height=44; r+=1
rh('A12',"Critical flags (evidence required)")
hdrcell(rf,13,1,"Flag"); hdrcell(rf,13,2,"Severity"); hdrcell(rf,13,3,"Effect")
flag_desc=[("F1 Fabricated source/quotation","Central/Periph","Caps D2 = 0 (central) or ≤1 (peripheral); integrity-flagged."),
 ("F2 Misattribution","—","Caps D2 ≤1; integrity-flagged if material."),
 ("F3 Stale-as-current","Central/Periph","Caps D3 = 0 (central) or ≤1 (peripheral); integrity-flagged."),
 ("F4 Invented obligation","Central/Periph","Caps D3 = 0 (central) or ≤1 (peripheral); integrity-flagged."),
 ("F5 False exhaustiveness","—","Caps D6 ≤1."),
 ("F6 Material scope evasion","—","Caps D1 ≤1; scope-flagged."),
 ("F7 Unsupported central legal conclusion","—","Caps D2 ≤1; integrity-flagged.")]
r=14
for a,s,b in flag_desc:
    rf.cell(row=r,column=1,value=a).font=font(10,True); rf.cell(row=r,column=1).alignment=wrap; rf.cell(row=r,column=1).border=border
    rf.cell(row=r,column=2,value=s).font=font(9); rf.cell(row=r,column=2).alignment=center; rf.cell(row=r,column=2).border=border
    rf.cell(row=r,column=3,value=b).font=font(10); rf.cell(row=r,column=3).alignment=wrap; rf.cell(row=r,column=3).border=border
    rf.row_dimensions[r].height=26; r+=1
rh('A22',"Tiers & math")
def kv(r,k,v):
    rf.cell(row=r,column=1,value=k).font=font(10,True); rf.cell(row=r,column=1).alignment=wrap
    rf.merge_cells(start_row=r,start_column=3,end_row=r,end_column=3)
    rf.cell(row=r,column=3,value=v).font=font(10); rf.cell(row=r,column=3).alignment=wrap
kv(23,"Overall","Sum over dimensions of weight x effective / 4 (weights total 100; perfect 4s = 100).")
kv(24,"Tiers","A 85-100 - B 70-84 - C 50-69 - D 30-49 - E 0-29.")
kv(25,"Reliability overlay","Report tier AND overlay. Any F1-F5/F7 -> INTEGRITY-FLAGGED; F6 -> SCOPE-FLAGGED. Overlay governs reliance regardless of score.")
for rr in (23,24,25): rf.row_dimensions[rr].height=30

wb.save(OUT)
print("saved", OUT)
